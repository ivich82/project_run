from wsgiref import headers

from django.contrib.sessions import serializers
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.conf import settings
from rest_framework import viewsets
from rest_framework.pagination import PageNumberPagination
from .models import Run, AthleteInfo, Challenge, Position, CollectibleItem
from .serializer import RunSerializer, UserSerializer, AthleteInfoSerializer, ChallengeSerializer, PositionSerializer, CollectibleItemSerializer, UserDetailSerializer
from django.contrib.auth.models import User
from rest_framework.filters import  SearchFilter
from rest_framework.views import APIView
from rest_framework import status
from django.shortcuts import get_object_or_404
from django.db.models import Sum, Max, Min, Count, Q, Avg
from geopy.distance import geodesic
from openpyxl import load_workbook
from datetime import datetime


@api_view(['GET'])
def company_details(request):
    details = {'company_name': settings.COMPANY_NAME,
               'slogan': settings.SLOGAN,
               'contacts': settings.CONTACTS}
    return Response(details)

class RunPagination(PageNumberPagination):
    page_size_query_param = 'size'
    max_page_size = 50

class RunViewSet(viewsets.ModelViewSet):
    queryset = Run.objects.select_related('athlete').all()
    serializer_class = RunSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ['status', 'athlete']
    ordering_fields = ['created_at']
    pagination_class = RunPagination


class UserPagination(PageNumberPagination):
    page_size_query_param = 'size'
    max_page_size = 50

class UserViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = User.objects.filter(is_superuser=False).annotate(
                runs_finished=Count('run', filter=Q(run__status='finished')))
    # queryset = User.objects.filter(is_superuser=False)
    serializer_class = UserSerializer
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['first_name', 'last_name']
    ordering_fields = ['date_joined']
    pagination_class = UserPagination

    def get_queryset(self):
        qs = self.queryset
        type = self.request.query_params.get('type', None)
        if type == 'coach':
            qs = qs.filter(is_staff=True)
        elif type == 'athlete':
            qs = qs.filter(is_staff=False)
        elif self.action == 'retrieve':
            qs = User.objects.prefetch_related('collectibleitems').filter(is_superuser=False)

        return qs

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return UserDetailSerializer
        return super().get_serializer_class()


class StartAPIView(APIView):

    def post(self, request, id):
        run = get_object_or_404(Run, id=id)
        if run.status == 'init':
            run.status = 'in_progress'
            run.save()
            data = {'status': 'in_progress'}
            return Response(data)
        return Response(status=400)


class StopAPIView(APIView):

    def post(self, request, id):
        run = get_object_or_404(Run, id=id)
        if run.status == 'in_progress':

            run_pos = Position.objects.filter(run=id).order_by('date_time')
            loc = list(map(lambda x: (x.latitude, x.longitude), run_pos))
            run.distance = sum(geodesic(loc[i - 1], loc[i]).km for i in range(1, len(loc)))

            run_times_pos = run_pos.aggregate(
                max_date_time = Max('date_time'),
                min_date_time = Min('date_time'),
                average_speed = Avg('speed')
            )
            min, max = run_times_pos['min_date_time'], run_times_pos['max_date_time']
            if max and min:
                seconds = (run_times_pos['max_date_time'] - run_times_pos['min_date_time']).total_seconds()
                run.run_time_seconds = seconds

            run.speed = run_times_pos['average_speed']
            run.status = 'finished'
            run.save()


            athlete = get_object_or_404(User, id=run.athlete.id)
            serializer = UserSerializer(athlete)
            count_run = serializer.data.get('runs_finished')
            if int(count_run) == 10:
                object, created = Challenge.objects.update_or_create(
                    athlete=run.athlete,
                    full_name='Сделай 10 Забегов!')

            if run.distance >= 2 and run.run_time_seconds <= 10 * 60:
                object, created = Challenge.objects.update_or_create(
                    athlete=run.athlete,
                    full_name='2 километра за 10 минут!')


            sum_distance = Run.objects.filter(status='finished', athlete=athlete.id).aggregate(Sum('distance'))
            if sum_distance['distance__sum'] and sum_distance['distance__sum'] >= 50:
                object, created = Challenge.objects.update_or_create(
                    athlete = run.athlete,
                    full_name = 'Пробеги 50 километров!')

            items = CollectibleItem.objects.all()
            for item in items:
                for i in range(0, len(loc) - 1):
                    if geodesic(loc[i], (item.latitude, item.longitude)).m <= 100:
                        athlete.collectibleitems.add(item)

            return Response({'status': 'finished'})
        return Response(status=400)


class AthleteInfoAPIView(APIView):

    def get(self, request, pk, format=None):
        user_obj = get_object_or_404(User, id=pk)
        object, created = AthleteInfo.objects.get_or_create(user_id=user_obj)
        serializer = AthleteInfoSerializer(object)
        return Response(serializer.data)

    def put(self, request, pk, format=None):
        user_obj = get_object_or_404(User, id=pk)
        object, created = AthleteInfo.objects.update_or_create(user_id=user_obj)
        serializer = AthleteInfoSerializer(object, data=request.data, partial=True)
        if serializer.is_valid(raise_exception=True):
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class ChallengeViewSet(viewsets.ModelViewSet):
    queryset = Challenge.objects.select_related('athlete').all()
    serializer_class = ChallengeSerializer

    def get_queryset(self):
        qs = self.queryset
        athlete = self.request.query_params.get('athlete', None)
        if athlete:
            qs = qs.filter(athlete=athlete)
        return  qs

class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.all()
    serializer_class = PositionSerializer

    def get_queryset(self):
        qs = self.queryset
        run = self.request.query_params.get('run', None)
        if run:
            qs = qs.filter(run=run)
        return  qs

    def perform_create(self, serializer):
        run = serializer.validated_data['run']
        queryset = Position.objects.filter(run=run).order_by('date_time')
        if queryset.exists():

            pos_lst = list(queryset)
            date_time = serializer.validated_data['date_time']
            latitude = serializer.validated_data['latitude']
            longitude = serializer.validated_data['longitude']

            distance_part = geodesic((latitude, longitude), (pos_lst[-1].latitude, pos_lst[-1].longitude)).meters
            seconds = (date_time - pos_lst[-1].date_time).total_seconds()

            speed = distance_part / seconds
            distance = pos_lst[-1].distance + distance_part / 1000

            serializer.save(distance=distance, speed=speed)
        else:
            serializer.save(distance=0, speed=0)

class  CollectibleItemViewSet(viewsets.ModelViewSet):
    queryset = CollectibleItem.objects.all()
    serializer_class = CollectibleItemSerializer

@api_view(['POST'])
def upload_file(request):
    file = request.FILES.get('file')
    if file:
        workbook = load_workbook(filename=file)
        sheet = workbook.active
        filedata = (row for row in sheet.iter_rows(values_only=True) if all(row) != False)
        lst = []
        for index, row in enumerate(filedata):
            if index > 0:
                incoming_data = {
                    "name": row[0],
                    "uid": row[1],
                    "latitude": row[3],
                    "longitude": row[4],
                    "picture": row[5],
                    "value": row[2]
                }
                serializer = CollectibleItemSerializer(data=incoming_data)
                if serializer.is_valid():
                    serializer.save()
                else:
                    lst.append(list(row))
        return Response(lst)
    return Response(status=status.HTTP_400_BAD_REQUEST)